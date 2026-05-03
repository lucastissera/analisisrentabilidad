"""
excel_engine.py — Export / Import / Template para el Sistema de Rentabilidad
"""

import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

from ipc_engine import adjust_monthly_store

# ─── Palette ──────────────────────────────────────────────────────────────────
P = {
    "navy":    "1A2744",
    "mid":     "1E3A5F",
    "accent":  "2563EB",
    "green":   "16A34A",
    "red":     "DC2626",
    "amber":   "D97706",
    "purple":  "7C3AED",
    "light":   "EBF2FB",
    "white":   "FFFFFF",
    "yellow":  "FEF9C3",
    "border":  "CBD5E1",
    "blue_lt": "DBEAFE",
    "green_lt":"D1FAE5",
    "red_lt":  "FEE2E2",
    "amber_lt":"FEF3C7",
    "gray":    "F1F5F9",
}

MONTHS_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

FIELD_LABELS = {
    "ventasBrutas":  "Ventas brutas ($)",
    "devoluciones":  "Devoluciones ($)",
    "otrosIngresos": "Otros ingresos ($)",
    "cmv":           "CMV / Mercadería ($)",
    "moDirecta":     "Mano de obra directa ($)",
    "materiales":    "Materiales e insumos ($)",
    "alquiler":      "Alquiler ($)",
    "sueldos":       "Sueldos y cargas ($)",
    "servicios":     "Servicios ($)",
    "mantenimiento": "Mantenimiento ($)",
    "marketing":     "Marketing local ($)",
    "admGeneral":    "Adm. general ($)",
    "inventario":    "Inventario promedio ($)",
    "equipamiento":  "Equipamiento ($)",
    "mobiliario":    "Mobiliario ($)",
    "empleados":     "Nº empleados",
    "m2":            "M² del local",
    "transacciones": "Transacciones/mes",
}
DATA_KEYS = list(FIELD_LABELS.keys())


def corp_costs_total(corp):
    """Suma gastos corporativos sin duplicar marketing_corp / marketing."""
    if not corp:
        return 0.0
    c = dict(corp)
    sub = 0.0
    for k, v in c.items():
        if k in ("marketing", "marketing_corp"):
            continue
        if isinstance(v, (int, float)):
            sub += float(v)
    if c.get("marketing_corp") is not None:
        sub += float(c.get("marketing_corp") or 0)
    elif c.get("marketing") is not None:
        sub += float(c.get("marketing") or 0)
    return sub


CORP_LABELS = {
    "gerencia":      "Gerencia y administración",
    "contabilidad":  "Contabilidad y auditoría",
    "sistemas":      "Sistemas y tecnología",
    "legal":         "Legal y compliance",
    "marketing_corp":"Marketing corporativo",
    "logistica":     "Logística central",
    "seguros":       "Seguros",
    "otros":         "Otros gastos centrales",
}

# ─── Style helpers ────────────────────────────────────────────────────────────
def _tb(color=None):
    c = color or P["border"]
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, name="Calibri", italic=False):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)

def _align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _cell(ws, row, col, value=None, bold=False, color="000000", size=10,
          fill=None, align_h="left", num_fmt=None, indent=0, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, color=color, size=size, italic=italic)
    c.alignment = _align(h=align_h, indent=indent)
    c.border = _tb()
    if fill:
        c.fill = _fill(fill)
    if num_fmt:
        c.number_format = num_fmt
    return c

def _header_row(ws, row, cols_values, bg=P["navy"], fg="FFFFFF", height=22, bold=True, size=10):
    ws.row_dimensions[row].height = height
    for col, val in cols_values:
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=bold, color=fg, size=size)
        c.fill = _fill(bg)
        c.alignment = _align(h="center")
        c.border = _tb()

def _section(ws, row, col_start, col_end, text, bg=P["mid"], fg="C9DEFF"):
    ws.row_dimensions[row].height = 16
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=f"  {text}")
    c.font = _font(bold=True, color=fg, size=9)
    c.fill = _fill(bg)
    c.alignment = _align(h="left")
    c.border = _tb()

def _merge_title(ws, row, c1, c2, text, bg=P["navy"], fg="FFFFFF", size=12, height=32):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.font = _font(bold=True, color=fg, size=size)
    c.fill = _fill(bg)
    c.alignment = _align(h="center")
    c.border = _tb()

def _set_col_widths(ws, widths):  # {col_idx: width}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

NUM_FMT   = '#.##0;(#.##0);-'
PCT_FMT   = '0,0%;(0,0%);-'
INPUT_FMT = '#.##0'

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ═══════════════════════════════════════════════════════════════════════════════
class ExcelExporter:
    def __init__(self, cfg, monthly_data):
        self.cfg = cfg
        self.monthly = (
            adjust_monthly_store(dict(monthly_data), cfg)
            if cfg.get("ipcAdjust")
            else dict(monthly_data)
        )
        self.branch_names = cfg.get("branchNames", [])
        self.empresa = cfg.get("empresa", "Empresa")
        self.period_type = cfg.get("periodType", "monthly")
        self.alloc_method = cfg.get("allocMethod", "ventas")
        self.corp = cfg.get("corpCosts", {})
        self.manual_pct = cfg.get("manualPct", [])

    def build(self) -> io.BytesIO:
        n = len(self.branch_names)
        if n == 0:
            wb = Workbook()
            ws = wb.active
            ws.title = "Aviso"
            ws["A1"] = "Configurá al menos una sucursal en el sistema para exportar el Excel completo."
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        wb = Workbook()
        self._sheet_portada(wb)
        self._sheet_datos_mensuales(wb)
        self._sheet_acumulado(wb)
        self._sheet_metricas(wb)
        self._sheet_corp(wb)

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Portada ───────────────────────────────────────────────────────────────
    def _sheet_portada(self, wb):
        ws = wb.create_sheet("📋 Portada")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1:4, 2:40, 3:25})

        _merge_title(ws, 1, 2, 3, f"ANÁLISIS DE RENTABILIDAD — {self.empresa.upper()}", height=40, size=14)
        _merge_title(ws, 2, 2, 3, f"Período: {self.cfg.get('period','')}", bg=P["mid"], size=11, height=24)

        mes_cierre = int(self.cfg.get("fiscalYearEndMonth", 12) or 12)
        meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ]
        fy_label = meses[mes_cierre - 1] if 1 <= mes_cierre <= 12 else str(mes_cierre)
        info = [
            ("Empresa:", self.empresa),
            ("Período:", self.cfg.get("period", "")),
            ("Cierre de ejercicio (mes):", fy_label),
            ("Tipo de análisis:", "Mensual" if self.period_type == "monthly" else "Anual"),
            ("Sucursales:", ", ".join(self.branch_names)),
            ("Método de prorrateo:", self.alloc_method.capitalize()),
            ("Ajuste IPC:", "Sí (datos exportados al cierre)" if self.cfg.get("ipcAdjust") else "No"),
            ("Gastos corporativos totales:", f"${corp_costs_total(self.corp):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")),
            ("Períodos cargados:", str(len(self.monthly))),
            ("Exportado:", __import__("datetime").datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]
        r = 4
        for label, val in info:
            ws.row_dimensions[r].height = 18
            _cell(ws, r, 2, label, bold=True, fill=P["light"])
            _cell(ws, r, 3, val, fill=P["white"])
            r += 1

        # Legend
        r += 1
        _merge_title(ws, r, 2, 3, "LEYENDA DE COLORES", bg=P["mid"], size=10, height=20); r+=1
        legend = [
            (P["yellow"], "Azul / Amarillo — Inputs del usuario"),
            (P["green_lt"], "Verde — Margen Bruto / EBITDA positivo"),
            (P["red_lt"], "Rojo — Valores negativos o críticos"),
            (P["blue_lt"], "Azul claro — Totales y consolidados"),
            (P["amber_lt"], "Amarillo — Indicadores de atención"),
        ]
        for fill_c, desc in legend:
            ws.row_dimensions[r].height = 16
            _cell(ws, r, 2, "   " + desc, fill=fill_c)
            _cell(ws, r, 3, "", fill=fill_c)
            r += 1

    # ── Datos mensuales ───────────────────────────────────────────────────────
    def _sheet_datos_mensuales(self, wb):
        ws = wb.create_sheet("📊 Datos por Período")
        ws.sheet_view.showGridLines = False
        n = len(self.branch_names)
        periods = sorted(self.monthly.keys())

        # Build columns: col1=label, then for each period N branch columns
        col_label = 2
        # col_start for period i: 3 + i*n
        total_cols = 2 + len(periods) * n + 1  # +1 for total

        _set_col_widths(ws, {1: 3, 2: 34})
        for c in range(3, total_cols + 2):
            ws.column_dimensions[get_column_letter(c)].width = 16

        # Row 1 — main title
        _merge_title(ws, 1, 2, 2 + len(periods)*n, f"DATOS POR PERÍODO — {self.empresa}", size=13)
        ws.freeze_panes = "C4"

        # Row 2 — period group headers
        ws.row_dimensions[2].height = 20
        _cell(ws, 2, 2, "CONCEPTO", bold=True, fill=P["navy"], color="FFFFFF", align_h="center")
        for p_idx, period in enumerate(periods):
            col_start = 3 + p_idx * n
            ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start + n - 1)
            label = self._period_label(period)
            c = ws.cell(row=2, column=col_start, value=label)
            c.font = _font(bold=True, color="FFFFFF", size=10)
            c.fill = _fill(P["mid"])
            c.alignment = _align(h="center")
            c.border = _tb()

        # Row 3 — branch sub-headers
        ws.row_dimensions[3].height = 18
        _cell(ws, 3, 2, "", fill=P["navy"])
        for p_idx, period in enumerate(periods):
            for b_idx, bname in enumerate(self.branch_names):
                col = 3 + p_idx * n + b_idx
                c = ws.cell(row=3, column=col, value=bname)
                c.font = _font(bold=True, color="FFFFFF", size=9)
                c.fill = _fill(P["accent"])
                c.alignment = _align(h="center")
                c.border = _tb()

        # Data rows
        sections = [
            ("INGRESOS", ["ventasBrutas","devoluciones","otrosIngresos"]),
            ("COSTO DE VENTAS", ["cmv","moDirecta","materiales"]),
            ("GASTOS OPERATIVOS", ["alquiler","sueldos","servicios","mantenimiento","marketing","admGeneral"]),
            ("ACTIVOS", ["inventario","equipamiento","mobiliario"]),
            ("DATOS OPERATIVOS", ["empleados","m2","transacciones"]),
        ]

        r = 4
        for sec_name, keys in sections:
            _section(ws, r, 2, 2 + len(periods)*n, sec_name); r+=1
            for key in keys:
                ws.row_dimensions[r].height = 16
                bg = P["light"] if r % 2 == 0 else P["white"]
                _cell(ws, r, 2, FIELD_LABELS[key], indent=1, fill=bg)
                for p_idx, period in enumerate(periods):
                    branches = self.monthly.get(period, [])
                    for b_idx in range(len(self.branch_names)):
                        col = 3 + p_idx * n + b_idx
                        val = branches[b_idx].get(key, 0) if b_idx < len(branches) else 0
                        c = ws.cell(row=r, column=col, value=val)
                        c.font = _font(color="0000FF")
                        c.fill = _fill(P["yellow"])
                        c.number_format = INPUT_FMT
                        c.alignment = _align(h="right")
                        c.border = _tb()
                r += 1

    # ── Acumulado ─────────────────────────────────────────────────────────────
    def _sheet_acumulado(self, wb):
        ws = wb.create_sheet("📈 Acumulado")
        ws.sheet_view.showGridLines = False
        n = len(self.branch_names)
        periods = sorted(self.monthly.keys())
        if not periods:
            _merge_title(ws, 1, 1, 2, "Sin datos cargados"); return

        _set_col_widths(ws, {1:3, 2:34, **{3+i:16 for i in range(n)}, 3+n:18})

        _merge_title(ws, 1, 2, 3+n, f"ACUMULADO — {self.empresa}", size=13)

        # Branch headers
        ws.row_dimensions[2].height = 20
        _cell(ws, 2, 2, "CONCEPTO", bold=True, fill=P["navy"], color="FFFFFF", align_h="center")
        for i, bname in enumerate(self.branch_names):
            _header_row(ws, 2, [(3+i, bname)], bg=P["mid"])
        _header_row(ws, 2, [(3+n, "TOTAL EMPRESA")], bg=P["accent"])

        # Aggregate all periods
        from collections import defaultdict
        acc = [{} for _ in range(n)]
        for period in periods:
            branches = self.monthly.get(period, [])
            for i, b in enumerate(branches):
                if i >= n: break
                for k, v in b.items():
                    if isinstance(v, (int, float)):
                        acc[i][k] = acc[i].get(k, 0) + v
                    else:
                        acc[i][k] = v

        def write_row(ws, r, label, key, fmt=NUM_FMT, bold=False, bg_label=None):
            ws.row_dimensions[r].height = 16
            bg = bg_label or (P["light"] if r % 2 == 0 else P["white"])
            _cell(ws, r, 2, label, bold=bold, fill=bg, indent=1 if not bold else 0)
            sum_f_parts = []
            for i in range(n):
                col = 3 + i
                cl = get_column_letter(col)
                val = acc[i].get(key, 0)
                c = ws.cell(row=r, column=col, value=val)
                c.font = _font(bold=bold)
                c.fill = _fill(bg)
                c.number_format = fmt
                c.alignment = _align(h="right")
                c.border = _tb()
                sum_f_parts.append(f"{cl}{r}")
            tot_col = 3 + n
            tot_c = ws.cell(row=r, column=tot_col, value=f"=SUM({','.join(sum_f_parts)})")
            tot_c.font = _font(bold=bold)
            tot_c.fill = _fill(P["blue_lt"])
            tot_c.number_format = fmt
            tot_c.alignment = _align(h="right")
            tot_c.border = _tb()
            return r + 1

        r = 3
        r = write_row(ws, r, "Ventas Netas ($)", "ventasBrutas", bold=True, bg_label=P["blue_lt"])
        _section(ws, r, 2, 3+n, "MARGEN BRUTO"); r+=1
        r = write_row(ws, r, "CMV Total ($)", "cmv")
        # Compute derived metrics manually for accumulation display
        for i in range(n):
            acc[i]["_vn"] = acc[i].get("ventasBrutas",0) - acc[i].get("devoluciones",0) + acc[i].get("otrosIngresos",0)
            acc[i]["_cmv_t"] = acc[i].get("cmv",0) + acc[i].get("moDirecta",0) + acc[i].get("materiales",0)
            acc[i]["_mb"] = acc[i]["_vn"] - acc[i]["_cmv_t"]
            acc[i]["_go"] = sum(acc[i].get(k,0) for k in ["alquiler","sueldos","servicios","mantenimiento","marketing","admGeneral"])
            acc[i]["_ebitda_d"] = acc[i]["_mb"] - acc[i]["_go"]

        r = write_row(ws, r, "Margen Bruto ($)", "_mb", bold=True, bg_label=P["green_lt"])
        _section(ws, r, 2, 3+n, "EBITDA"); r+=1
        r = write_row(ws, r, "Gastos Operativos Directos ($)", "_go")
        r = write_row(ws, r, "EBITDA Directo ($)", "_ebitda_d", bold=True, bg_label=P["green_lt"])

        # Periods summary
        r += 1
        _section(ws, r, 2, 3+n, f"PERÍODOS INCLUIDOS: {len(periods)} — {', '.join(self._period_label(p) for p in periods)}"); r+=1

    # ── Métricas / Dashboard ──────────────────────────────────────────────────
    def _sheet_metricas(self, wb):
        ws = wb.create_sheet("🎯 Métricas y KPIs")
        ws.sheet_view.showGridLines = False
        n = len(self.branch_names)
        periods = sorted(self.monthly.keys())

        _set_col_widths(ws, {1:3, 2:34, **{3+i:16 for i in range(n)}, 3+n:18})
        _merge_title(ws, 1, 2, 3+n, f"MÉTRICAS Y KPIs — {self.empresa}", size=13)

        ws.row_dimensions[2].height = 20
        _cell(ws, 2, 2, "KPI", bold=True, fill=P["navy"], color="FFFFFF", align_h="center")
        for i, bname in enumerate(self.branch_names):
            _header_row(ws, 2, [(3+i, bname)], bg=P["mid"])
        _header_row(ws, 2, [(3+n, "EMPRESA")], bg=P["accent"])

        # For the last loaded period (or cumulative if multiple)
        if not periods:
            return
        # Use cumulative if multiple periods
        if len(periods) > 1:
            from collections import defaultdict
            acc = [{} for _ in range(n)]
            for period in periods:
                for i, b in enumerate(self.monthly.get(period, [])):
                    if i >= n: break
                    for k, v in b.items():
                        if isinstance(v, (int, float)):
                            acc[i][k] = acc[i].get(k, 0) + v
                        else:
                            acc[i][k] = v
            branches_for_metrics = acc
        else:
            branches_for_metrics = self.monthly[periods[0]]

        # Compute
        corp_total = corp_costs_total(self.corp)
        weights = self._alloc_w(branches_for_metrics)
        metrics = []
        for i, b in enumerate(branches_for_metrics[:n]):
            vn = b.get("ventasBrutas",0) - b.get("devoluciones",0) + b.get("otrosIngresos",0)
            cmv = b.get("cmv",0) + b.get("moDirecta",0) + b.get("materiales",0)
            mb = vn - cmv
            go = sum(b.get(k,0) for k in ["alquiler","sueldos","servicios","mantenimiento","marketing","admGeneral"])
            corp_share = corp_total * weights[i]
            ebitda_d = mb - go
            ebitda_n = ebitda_d - corp_share
            activos = b.get("inventario",0) + b.get("equipamiento",0) + b.get("mobiliario",0)
            metrics.append({
                "Ventas Netas ($)": vn,
                "CMV ($)": cmv,
                "Margen Bruto ($)": mb,
                "% Margen Bruto": mb/vn if vn else 0,
                "Gastos Op. ($)": go,
                "EBITDA Directo ($)": ebitda_d,
                "% EBITDA Directo": ebitda_d/vn if vn else 0,
                "Corp. Asignado ($)": corp_share,
                "EBITDA Neto ($)": ebitda_n,
                "% EBITDA Neto": ebitda_n/vn if vn else 0,
                "Total Activos ($)": activos,
                "ROA": ebitda_n/activos if activos else 0,
                "Punto de Equilibrio ($)": self._pe(b, vn, cmv),
                "Venta por Empleado ($)": vn/b["empleados"] if b.get("empleados") else 0,
                "Venta por M² ($)": vn/b["m2"] if b.get("m2") else 0,
                "Ticket Promedio ($)": vn/b["transacciones"] if b.get("transacciones") else 0,
            })

        pct_keys = {"% Margen Bruto","% EBITDA Directo","% EBITDA Neto","ROA"}
        bold_keys = {"Ventas Netas ($)","Margen Bruto ($)","EBITDA Directo ($)","EBITDA Neto ($)"}
        section_before = {
            "Margen Bruto ($)": "RENTABILIDAD BRUTA",
            "EBITDA Directo ($)": "EBITDA",
            "Corp. Asignado ($)": "GASTOS CORPORATIVOS",
            "EBITDA Neto ($)": "RESULTADO NETO",
            "Total Activos ($)": "EFICIENCIA DE ACTIVOS",
            "Punto de Equilibrio ($)": "PUNTO DE EQUILIBRIO",
            "Venta por Empleado ($)": "INDICADORES OPERATIVOS",
        }

        if not metrics:
            _merge_title(ws, 3, 2, 3 + n, "No hay sucursales con datos para calcular métricas.")
            return
        r = 3
        for kpi_label, _ in metrics[0].items():
            if kpi_label in section_before:
                _section(ws, r, 2, 3+n, section_before[kpi_label]); r+=1
            ws.row_dimensions[r].height = 16
            is_bold = kpi_label in bold_keys
            is_pct = kpi_label in pct_keys
            bg = P["green_lt"] if is_bold else (P["light"] if r%2==0 else P["white"])
            fmt = PCT_FMT if is_pct else NUM_FMT
            _cell(ws, r, 2, kpi_label, bold=is_bold, fill=bg, indent=0 if is_bold else 1)
            vals = [m.get(kpi_label, 0) for m in metrics]
            for i, v in enumerate(vals):
                col = 3 + i
                c = ws.cell(row=r, column=col, value=v)
                color = "000000"
                if is_pct:
                    color = P["green"] if v >= (0.35 if "Margen" in kpi_label else 0.15 if "EBITDA" in kpi_label else 0.10) else (P["amber"] if v > 0 else P["red"])
                c.font = _font(bold=is_bold, color=color)
                c.fill = _fill(bg)
                c.number_format = fmt
                c.alignment = _align(h="right")
                c.border = _tb()
            # Total col
            tot_col = 3 + n
            if is_pct:
                tot_vn = sum(m.get("Ventas Netas ($)",0) for m in metrics)
                nom_key = kpi_label.replace("% ","") + " ($)" if "% " in kpi_label else None
                if nom_key:
                    nom_val = sum(m.get(nom_key,0) for m in metrics)
                    tot_val = nom_val / tot_vn if tot_vn else 0
                else:
                    tot_val = sum(vals)/len(vals) if vals else 0
            else:
                tot_val = sum(vals)
            tc = ws.cell(row=r, column=tot_col, value=tot_val)
            tc.font = _font(bold=is_bold)
            tc.fill = _fill(P["blue_lt"])
            tc.number_format = fmt
            tc.alignment = _align(h="right")
            tc.border = _tb()
            r += 1

    # ── Gastos corporativos ───────────────────────────────────────────────────
    def _sheet_corp(self, wb):
        ws = wb.create_sheet("🏢 Corporativos")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1:3, 2:38, 3:20, 4:20})

        _merge_title(ws, 1, 2, 4, "GASTOS CORPORATIVOS Y PRORRATEO", size=12)
        _header_row(ws, 2, [(2,"CONCEPTO"),(3,"MONTO ($)"),(4,"NOTAS")], bg=P["mid"])

        r = 3
        corp_keys = list(CORP_LABELS.keys())
        for k in corp_keys:
            val = self.corp.get(k, self.corp.get(k.replace("_corp",""), 0))
            bg = P["light"] if r%2==0 else P["white"]
            ws.row_dimensions[r].height = 16
            _cell(ws, r, 2, CORP_LABELS[k], fill=bg)
            c = ws.cell(row=r, column=3, value=val)
            c.font = _font(color="0000FF")
            c.fill = _fill(P["yellow"])
            c.number_format = NUM_FMT
            c.alignment = _align(h="right")
            c.border = _tb()
            _cell(ws, r, 4, "", fill=bg)
            r += 1

        # Total
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 2, "TOTAL GASTOS CORPORATIVOS", bold=True, fill=P["blue_lt"])
        c = ws.cell(row=r, column=3, value=f"=SUM(C3:C{r-1})")
        c.font = _font(bold=True)
        c.fill = _fill(P["blue_lt"])
        c.number_format = NUM_FMT
        c.alignment = _align(h="right")
        c.border = _tb()
        _cell(ws, r, 4, "", fill=P["blue_lt"])
        r += 2

        # Alloc method
        _section(ws, r, 2, 4, "MÉTODO DE PRORRATEO ACTIVO"); r+=1
        ws.row_dimensions[r].height = 18
        alloc_labels = {
            "ventas":"Por % de Ventas","empleados":"Por Nº de Empleados",
            "m2":"Por M² del Local","igualitario":"Igualitario",
            "transacciones":"Por Transacciones","manual":"Manual / Mixto"
        }
        _cell(ws, r, 2, "Método seleccionado:", bold=True, fill=P["light"])
        c = ws.cell(row=r, column=3, value=alloc_labels.get(self.alloc_method, self.alloc_method))
        c.font = _font(bold=True, color=P["accent"])
        c.fill = _fill(P["yellow"])
        c.border = _tb()
        _cell(ws, r, 4, "", fill=P["light"])
        r += 1

        # Manual pct table
        if self.alloc_method == "manual" and self.manual_pct:
            r += 1
            _section(ws, r, 2, 4, "PORCENTAJES MANUALES"); r+=1
            for i, (bname, pct) in enumerate(zip(self.branch_names, self.manual_pct)):
                ws.row_dimensions[r].height = 16
                bg = P["light"] if r%2==0 else P["white"]
                _cell(ws, r, 2, bname, fill=bg)
                c = ws.cell(row=r, column=3, value=pct/100)
                c.font = _font(color="0000FF")
                c.fill = _fill(P["yellow"])
                c.number_format = PCT_FMT
                c.alignment = _align(h="right")
                c.border = _tb()
                _cell(ws, r, 4, "", fill=bg)
                r += 1

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _period_label(self, period):
        if len(period) == 7:  # "2024-01"
            try:
                y, m = period.split("-")
                return f"{MONTHS_ES[int(m)-1]} {y}"
            except:
                return period
        return period

    def _alloc_w(self, branches):
        n = len(branches)
        method = self.alloc_method
        if method == "ventas":
            w = [b.get("ventasBrutas",0)-b.get("devoluciones",0)+b.get("otrosIngresos",0) for b in branches]
        elif method == "empleados":
            w = [b.get("empleados",1) for b in branches]
        elif method == "m2":
            w = [b.get("m2",1) for b in branches]
        elif method == "transacciones":
            w = [b.get("transacciones",1) for b in branches]
        elif method == "manual":
            total = sum(self.manual_pct[:n]) or 1
            return [p/total for p in self.manual_pct[:n]]
        else:
            w = [1]*n
        total = sum(w) or 1
        return [x/total for x in w]

    def _pe(self, b, vn, cmv):
        cf = sum(b.get(k,0) for k in ["alquiler","sueldos","servicios","mantenimiento","admGeneral"])
        cv_pct = (cmv + b.get("marketing",0)) / vn if vn else 0
        return cf / (1-cv_pct) if (1-cv_pct) > 0 else 0


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL TEMPLATE GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════
class ExcelTemplateGenerator:
    """Generates a blank import template tailored to the current config."""

    def __init__(self, cfg):
        self.cfg = cfg
        self.branch_names = cfg.get("branchNames", [f"Sucursal {i+1}" for i in range(5)])
        self.empresa = cfg.get("empresa", "Empresa")
        self.period_type = cfg.get("periodType", "monthly")
        self.n = len(self.branch_names)

    def build(self) -> io.BytesIO:
        wb = Workbook()
        self._sheet_instrucciones(wb)
        self._sheet_template(wb)
        self._sheet_corporativos(wb)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _sheet_instrucciones(self, wb):
        ws = wb.create_sheet("📋 Instrucciones")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1:3, 2:55, 3:30})
        _merge_title(ws, 1, 2, 3, f"PLANTILLA DE IMPORTACIÓN — {self.empresa}", size=13, height=36)
        steps = [
            ("PASO 1 — Hoja 'Datos por Período'",
             "Complete las filas de datos para cada sucursal y período. Las celdas en AMARILLO son editables."),
            ("PASO 2 — Columna 'Período'",
             "Formato mensual: YYYY-MM (ej: 2024-01). Anual: YYYY (ej: 2024). Cada fila = un período + sucursal."),
            ("PASO 3 — Hoja 'Corporativos'",
             "Ingrese los gastos centrales anuales de la empresa."),
            ("PASO 4 — Importar en el sistema",
             "En el Dashboard use el botón 'Importar Excel' y seleccione este archivo guardado con sus datos."),
            ("COLORES",
             "🟡 Fondo amarillo = celda a completar  |  🔵 Texto azul = input  |  ⚫ Negro = fórmula automática"),
        ]
        r = 3
        for title, desc in steps:
            _merge_title(ws, r, 2, 3, title, bg=P["accent"], fg="FFFFFF", size=10, height=18); r+=1
            ws.row_dimensions[r].height = 32
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            c = ws.cell(row=r, column=2, value=desc)
            c.font = _font(size=10)
            c.alignment = _align(wrap=True)
            c.border = _tb()
            r += 1

    def _sheet_template(self, wb):
        ws = wb.create_sheet("📊 Datos por Período")
        ws.sheet_view.showGridLines = False

        # Columns: Periodo | Sucursal | field1 | field2 | ...
        headers = ["Período", "Sucursal"] + [FIELD_LABELS[k] for k in DATA_KEYS]
        _set_col_widths(ws, {1:3, 2:14, 3:24, **{4+i:15 for i in range(len(DATA_KEYS))}})

        _merge_title(ws, 1, 2, 2+len(headers)-1, "PLANTILLA DE IMPORTACIÓN DE DATOS", size=12)
        _header_row(ws, 2, [(2+i, h) for i, h in enumerate(headers)], bg=P["navy"], height=20)
        ws.freeze_panes = "D3"

        # Sample rows for each branch × 3 periods
        period_type = self.period_type
        sample_periods = ["2024-01","2024-02","2024-03"] if period_type=="monthly" else ["2024","2025","2026"]
        r = 3
        sample_vals = [0] * len(DATA_KEYS)
        for period in sample_periods:
            for b_idx, bname in enumerate(self.branch_names):
                ws.row_dimensions[r].height = 16
                # Period
                c = ws.cell(row=r, column=2, value=period)
                c.font = _font(color="0000FF"); c.fill = _fill(P["yellow"])
                c.alignment = _align(h="center"); c.border = _tb()
                # Branch
                c2 = ws.cell(row=r, column=3, value=bname)
                c2.font = _font(color="0000FF"); c2.fill = _fill(P["yellow"])
                c2.alignment = _align(h="left"); c2.border = _tb()
                # Data
                for k_idx, key in enumerate(DATA_KEYS):
                    col = 4 + k_idx
                    factor = 0.9**b_idx
                    val = round(sample_vals[k_idx] * factor)
                    c3 = ws.cell(row=r, column=col, value=val)
                    c3.font = _font(color="0000FF")
                    c3.fill = _fill(P["yellow"])
                    c3.number_format = INPUT_FMT
                    c3.alignment = _align(h="right")
                    c3.border = _tb()
                r += 1

        # Add empty rows for user to fill
        for _ in range(20):
            ws.row_dimensions[r].height = 16
            for col in range(2, 4+len(DATA_KEYS)):
                c = ws.cell(row=r, column=col, value="")
                c.fill = _fill(P["yellow"])
                c.border = _tb()
                if col >= 4:
                    c.number_format = INPUT_FMT
            r += 1

    def _sheet_corporativos(self, wb):
        ws = wb.create_sheet("🏢 Corporativos")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1:3, 2:38, 3:20})
        _merge_title(ws, 1, 2, 3, "GASTOS CORPORATIVOS ANUALES", size=12)
        _header_row(ws, 2, [(2,"CONCEPTO"),(3,"MONTO ($)")], bg=P["mid"])
        corp_defaults = {
            "gerencia": 0, "contabilidad": 0, "sistemas": 0,
            "legal": 0, "marketing_corp": 0, "logistica": 0,
            "seguros": 0, "otros": 0,
        }
        r = 3
        for k, label in CORP_LABELS.items():
            ws.row_dimensions[r].height = 16
            bg = P["light"] if r%2==0 else P["white"]
            _cell(ws, r, 2, label, fill=bg)
            c = ws.cell(row=r, column=3, value=corp_defaults.get(k,0))
            c.font = _font(color="0000FF")
            c.fill = _fill(P["yellow"])
            c.number_format = INPUT_FMT
            c.alignment = _align(h="right")
            c.border = _tb()
            r += 1
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 2, "TOTAL", bold=True, fill=P["blue_lt"])
        tc = ws.cell(row=r, column=3, value=f"=SUM(C3:C{r-1})")
        tc.font = _font(bold=True)
        tc.fill = _fill(P["blue_lt"])
        tc.number_format = NUM_FMT
        tc.alignment = _align(h="right")
        tc.border = _tb()


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL IMPORTER
# ═══════════════════════════════════════════════════════════════════════════════
class ExcelImporter:
    """Parses an uploaded Excel file (export or template format) back to JSON."""

    def __init__(self, stream):
        self.wb = load_workbook(stream, data_only=True)

    def parse(self) -> dict:
        result = {"monthly_data": {}, "config": None}

        # Try "Datos por Período" sheet
        if "📊 Datos por Período" in self.wb.sheetnames:
            result["monthly_data"] = self._parse_data_sheet()
        elif "Datos por Período" in self.wb.sheetnames:
            result["monthly_data"] = self._parse_data_sheet("Datos por Período")

        # Try corporate costs
        corp = self._parse_corp_sheet()
        if corp:
            result["config"] = {"corpCosts": corp}

        return result

    def _parse_data_sheet(self, sheet_name="📊 Datos por Período"):
        ws = self.wb[sheet_name]
        monthly = {}

        # Find header row (row with "Período" or "Periodo")
        header_row = None
        col_map = {}
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").strip()
                if val in ("Período", "Periodo", "Period"):
                    header_row = cell.row
                    break
            if header_row:
                break

        if not header_row:
            raise ValueError("No se encontró la fila de encabezados. Asegurate de usar la plantilla provista.")

        # Build column map from header row
        for cell in ws[header_row]:
            val = str(cell.value or "").strip()
            if val in ("Período", "Periodo"):
                col_map["__period__"] = cell.column
            elif val in ("Sucursal",):
                col_map["__branch__"] = cell.column
            else:
                # Match to DATA_KEYS by label
                for k, lbl in FIELD_LABELS.items():
                    if val.lower() == lbl.lower() or val.split("(")[0].strip().lower() == lbl.split("(")[0].strip().lower():
                        col_map[k] = cell.column
                        break

        if "__period__" not in col_map:
            raise ValueError("Columna 'Período' no encontrada.")

        # Parse data rows
        branch_order = []  # order in which branches appear for each period
        temp = {}  # {period: {branch_name: {field: val}}}

        for row_idx in range(header_row+1, ws.max_row+1):
            period_val = ws.cell(row=row_idx, column=col_map["__period__"]).value
            if not period_val:
                continue
            period = str(period_val).strip()
            branch_val = ws.cell(row=row_idx, column=col_map.get("__branch__", 1)).value
            branch = str(branch_val).strip() if branch_val else f"Sucursal {row_idx}"

            if period not in temp:
                temp[period] = {}
            if branch not in temp[period]:
                temp[period][branch] = {}

            for k, col in col_map.items():
                if k.startswith("__"):
                    continue
                cell_val = ws.cell(row=row_idx, column=col).value
                try:
                    temp[period][branch][k] = float(cell_val) if cell_val is not None else 0
                except (TypeError, ValueError):
                    temp[period][branch][k] = 0

        # Convert temp → {period: [branch_data, ...]}  (ordered by first appearance)
        branch_order_global = []
        for period_data in temp.values():
            for b in period_data:
                if b not in branch_order_global:
                    branch_order_global.append(b)

        for period, period_data in temp.items():
            monthly[period] = [period_data.get(b, {}) for b in branch_order_global]

        return monthly

    def _parse_corp_sheet(self):
        corp = {}
        for sheet_name in ["🏢 Corporativos", "Corporativos"]:
            if sheet_name not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet_name]
            reverse_labels = {v.lower(): k for k, v in CORP_LABELS.items()}
            # CONCEPTO en columna B, MONTO en C (igual que export/plantilla)
            for row in ws.iter_rows(min_row=3, min_col=2, max_col=3):
                label_cell = row[0]
                val_cell = row[1] if len(row) > 1 else None
                if not label_cell or not label_cell.value:
                    continue
                lbl = str(label_cell.value).strip().lower()
                if "total" in lbl and "gastos" in lbl:
                    continue
                if lbl == "total":
                    continue
                key = reverse_labels.get(lbl)
                if key and val_cell and val_cell.value is not None:
                    try:
                        corp[key] = float(val_cell.value)
                    except (TypeError, ValueError):
                        pass
            break
        return corp if corp else None
